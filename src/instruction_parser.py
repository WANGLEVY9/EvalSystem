"""
指令解析器 (v3.0)

核心改进:
1. 多通道解析 - JSON / Markdown / Excel / 自然语言
2. LLM 分段抽取 - 避免一次性 prompt 截断, 分别抽取检查点/约束/流程/知识 4 个独立 LLM call
3. 鲁棒的 Markdown 抽取 - 兼容更多变体: 中英文条件词、嵌套子步骤、## ### 多级标题
4. 集成 InstructionCompiler - 解析后自动编译 (复杂度/DAG/约束分类等)
5. Pydantic 校验 + 自动重试 - JSON 解析失败时降级到规则解析

支持格式:
- # Role / # Task / # Opening Line
- # Call Flow / # Conversation Flow / ## Step N / ### N.M
- # Knowledge Points / # Knowledge / # FAQ
- # Constraints / ## Constraints
"""

import json
import logging
import re
from pathlib import Path
from typing import Any, Optional

from openai import OpenAI

from .instruction_compiler import InstructionCompiler
from .models import (
    Checkpoint,
    CheckpointType,
    ConditionalBranch,
    Constraint,
    ConstraintSeverity,
    ExceptionRule,
    FlowNode,
    KnowledgeEntry,
    ScenarioTrigger,
    TaskInstruction,
    UtteranceConstraint,
    VariableDefinition,
)

logger = logging.getLogger(__name__)


# ============ LLM 分段抽取 Prompt ============

CHECKPOINTS_PROMPT = """你是对话任务分析专家。请从下列任务指令中抽取 **必须评估的检查点**。
检查点是被测系统在对话中应当完成的具体动作，必须可被外部观察到（不是心智活动）。

任务指令:
---
{instruction_text}
---

输出严格 JSON, 字段:
```json
{{"checkpoints":[
  {{"name":"...","description":"...","type":"must_do|should_do|optional",
    "keywords":["..."],"evaluation_criteria":"如何判断完成","order":1,
    "condition":"仅当条件满足时评估，例如：用户说在开车","at_flow_step":"关联步骤"}}
]}}
```

要求:
- 同一动作只列一次, 不要把流程节点 (步骤1/2) 直接拷成检查点
- 优先抽取 **必达类** (must_do): 如身份确认、关键信息告知、地址确认
- 条件性检查点 (如"若用户说不在家则改约") 务必填 condition 字段
- 只输出 JSON, 不要解释
"""

CONSTRAINTS_PROMPT = """你是对话合规分析专家。请从下列任务指令中抽取 **约束规则** 和 **话术约束**。

任务指令:
---
{instruction_text}
---

输出严格 JSON:
```json
{{
  "constraints":[
    {{"name":"...","description":"完整描述","severity":"critical|major|minor",
      "violation_examples":["..."],"check_method":"char_count|forbidden_words|regex|llm_judge"}}
  ],
  "utterance_constraints":[
    {{"name":"...","max_chars":30,"min_chars":null,"forbidden_words":["好的","哈哈"],
      "required_tone":"随意|简洁|正式","applies_to":"all|opening|closing","description":"..."}}
  ],
  "scenario_triggers":[
    {{"trigger_condition":"用户说在开车","trigger_keywords":["在开车","开车"],
      "expected_system_action":"礼貌说稍后再打并挂断",
      "forbidden_system_action":"继续推销",
      "should_end_conversation":true,"at_flow_step":""}}
  ]
}}
```

要求:
- 字数限制、禁词列表 → 必须放进 utterance_constraints
- 涉及"若用户...就..."的行为反应规则 → 放进 scenario_triggers
- 其他不能违反的规则 → 放进 constraints
- severity: critical=违反直接判定严重错误（如禁词、虚假承诺）；major=主要规则；minor=语气类
- 只输出 JSON
"""

FLOW_PROMPT = """你是对话流程分析专家。请从下列任务指令中抽取 **完整的对话流程节点**, 包括分支与子步骤。

任务指令:
---
{instruction_text}
---

输出严格 JSON:
```json
{{"flow_nodes":[
  {{"name":"步骤1: 身份确认","description":"...","step_number":1,
    "next_nodes":["步骤2: 通知主旨"],"is_start":true,"is_end":false,
    "conditions":{{"步骤2: 通知主旨":"用户确认身份"}},
    "branches":[
      {{"condition":"是负责人","condition_keywords":["是","对","我是"],
        "action":"进入步骤2","next_node":"步骤2: 通知主旨","is_default":false}},
      {{"condition":"不是负责人","condition_keywords":["不是","其他人"],
        "action":"请其转达后进入步骤2","next_node":"步骤2: 通知主旨","is_default":false}}
    ],
    "reference_script":"参考话术原文","sub_steps":["子操作1","子操作2"]}}
]}}
```

要求:
- 每个 ## Step N 或 N. xxxx 都对应一个 flow_node
- ### N.M (如 4.1 询问发布方式) 也独立成节点, 且 step_number=N.M
- branches 必须列出所有 if-else 分支 (含"否则/默认")
- 仅当 next_node 指向具体下一步时填写, 否则留空字符串
- 只输出 JSON
"""

KNOWLEDGE_PROMPT = """你是知识库分析专家。请从下列任务指令中抽取 **知识库条目** 和 **变量定义**。

任务指令:
---
{instruction_text}
---

输出严格 JSON:
```json
{{
  "knowledge_base":[
    {{"question":"用户可能问的问题","answer":"知识库标准答案",
      "keywords":["关键词"],"category":"general","priority":1}}
  ],
  "variables":[
    {{"name":"rider_name","placeholder":"${{rider_name}}","description":"骑手姓名",
      "sample_values":["张三","李四"],"default_value":"张师傅"}}
  ],
  "exception_rules":[
    {{"trigger":"被问及超出职责的问题","expected_behavior":"回复:我向同事确认后回电","forbidden_behavior":"乱编"}}
  ]
}}
```

要求:
- knowledge_base 既包括 # Knowledge Points 列表项, 也包括散落在流程里的关键事实(如"延迟约5-10秒")
- variables 必须包含所有 ${{...}} 和 **X 单** 形式的占位符
- exception_rules 来自约束中的"如被问.../如果用户..."模式
- 只输出 JSON
"""


# ============ 主解析器 ============


class InstructionParser:
    """任务指令解析器 v3 - 多通道 + 分段 LLM"""

    def __init__(
        self,
        llm_client: Optional[OpenAI] = None,
        model: str = "deepseek-chat",
        auto_compile: bool = True,
    ):
        self.llm_client = llm_client
        self.model = model
        self.auto_compile = auto_compile

    # ---- 入口 ----

    def parse_from_file(self, file_path: str | Path) -> TaskInstruction:
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"指令文件不存在: {file_path}")

        suffix = file_path.suffix.lower()

        if suffix == ".json":
            with open(file_path, encoding="utf-8") as f:
                data = json.load(f)
            instruction = self._parse_structured_data(data)
        elif suffix in (".md", ".markdown", ".txt"):
            with open(file_path, encoding="utf-8") as f:
                content = f.read()
            instruction = self.parse_from_markdown(content)
        elif suffix in (".xlsx", ".xls"):
            return self.parse_from_excel(file_path)[0]
        else:
            with open(file_path, encoding="utf-8") as f:
                content = f.read()
            try:
                data = json.loads(content)
                instruction = self._parse_structured_data(data)
            except json.JSONDecodeError:
                instruction = self.parse_from_markdown(content)

        if self.auto_compile:
            InstructionCompiler(instruction).compile()
        return instruction

    def parse_from_excel(self, file_path: str | Path) -> list[TaskInstruction]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        ws = wb.active
        instructions: list[TaskInstruction] = []
        for row_num in range(2, ws.max_row + 1):
            cell_id = ws.cell(row=row_num, column=1).value
            cell_content = ws.cell(row=row_num, column=2).value
            if not cell_content or not str(cell_content).strip():
                continue
            content = str(cell_content).strip()
            logger.info(f"解析 Excel 第{row_num}行 (ID={cell_id})...")
            try:
                instruction = self.parse_from_markdown(content)
                instruction.task_id = str(cell_id) if cell_id else instruction.task_id
                if self.auto_compile:
                    InstructionCompiler(instruction).compile()
                instructions.append(instruction)
            except Exception as e:
                logger.warning(f"解析第{row_num}行失败: {e}")

        logger.info(f"从 Excel 解析出 {len(instructions)} 个任务指令")
        return instructions

    def parse_from_markdown(self, markdown_text: str) -> TaskInstruction:
        sections = self._extract_markdown_sections(markdown_text)

        if self.llm_client:
            try:
                instruction = self._llm_parse_markdown_sectioned(markdown_text, sections)
                if self.auto_compile:
                    InstructionCompiler(instruction).compile()
                return instruction
            except Exception as e:
                logger.warning(f"LLM 分段解析失败, 回退规则解析: {e}")

        instruction = self._rule_parse_markdown(markdown_text, sections)
        if self.auto_compile:
            InstructionCompiler(instruction).compile()
        return instruction

    def parse_from_dict(self, data: dict[str, Any]) -> TaskInstruction:
        instr = self._parse_structured_data(data)
        if self.auto_compile:
            InstructionCompiler(instr).compile()
        return instr

    # ---- LLM 分段抽取 ----

    def _llm_parse_markdown_sectioned(
        self,
        markdown_text: str,
        sections: dict[str, str],
    ) -> TaskInstruction:
        """分段调用 LLM, 每次专注一个领域，避免输出截断"""
        logger.info("使用 LLM 分段抽取 (4 段并行)...")

        # 4 个独立 LLM 调用
        cp_data = self._call_llm_extract(CHECKPOINTS_PROMPT, markdown_text, "checkpoints", default={"checkpoints": []})
        con_data = self._call_llm_extract(CONSTRAINTS_PROMPT, markdown_text, "constraints", default={"constraints": [], "utterance_constraints": [], "scenario_triggers": []})
        flow_data = self._call_llm_extract(FLOW_PROMPT, markdown_text, "flow_nodes", default={"flow_nodes": []})
        kb_data = self._call_llm_extract(KNOWLEDGE_PROMPT, markdown_text, "knowledge_base", default={"knowledge_base": [], "variables": [], "exception_rules": []})

        # 组装
        role_desc = sections.get("role", "")
        task_desc = sections.get("task", "")
        # 重要: 保留原始变量占位符 (${var} / **X 单**), 由 dialogue_engine 在运行时替换为 default_value
        opening_line = sections.get("opening_line", "").strip()

        merged = {
            "task_name": self._infer_task_name(role_desc, task_desc),
            "task_description": task_desc or role_desc,
            "role_description": role_desc,
            "opening_line": opening_line,
            "system_prompt": self._build_system_prompt(role_desc, task_desc, sections),
            "checkpoints": cp_data.get("checkpoints", []),
            "constraints": con_data.get("constraints", []),
            "utterance_constraints": con_data.get("utterance_constraints", []),
            "scenario_triggers": con_data.get("scenario_triggers", []),
            "flow_nodes": flow_data.get("flow_nodes", []),
            "knowledge_base": kb_data.get("knowledge_base", []),
            "variables": kb_data.get("variables", []),
            "exception_rules": kb_data.get("exception_rules", []),
            "instruction_format": "markdown",
        }

        instruction = self._parse_structured_data(merged)
        instruction.raw_instruction = markdown_text
        return instruction

    def _call_llm_extract(
        self,
        prompt_template: str,
        instruction_text: str,
        primary_field: str,
        default: dict[str, Any],
        max_retries: int = 2,
    ) -> dict[str, Any]:
        """带重试的 LLM 抽取"""
        prompt = prompt_template.format(instruction_text=instruction_text)
        last_err: Optional[Exception] = None
        for attempt in range(max_retries + 1):
            try:
                response = self.llm_client.chat.completions.create(
                    model=self.model,
                    messages=[
                        {"role": "system", "content": "你是任务指令分析专家。严格按要求输出 JSON, 不要包含其他内容。"},
                        {"role": "user", "content": prompt},
                    ],
                    temperature=0.0,
                    max_tokens=4096,
                    response_format={"type": "json_object"},
                )
                text = response.choices[0].message.content
                data = self._parse_json(text)
                if primary_field in data or any(k in data for k in default.keys()):
                    return data
            except Exception as e:
                last_err = e
                logger.warning(f"LLM 抽取 {primary_field} 失败 (第{attempt + 1}次): {e}")
        if last_err:
            logger.warning(f"LLM 抽取 {primary_field} 全部重试失败, 使用默认值: {last_err}")
        return default

    # ---- 规则解析 (兜底) ----

    def _rule_parse_markdown(self, markdown_text: str, sections: dict[str, str]) -> TaskInstruction:
        role_desc = sections.get("role", "")
        task_desc = sections.get("task", "")
        # 重要: 保留原始变量占位符 (${var} / **X 单**), 由 dialogue_engine 在运行期替换为 default_value
        opening_line = sections.get("opening_line", "").strip()

        constraints = self._parse_constraints_from_text(sections.get("constraints", ""))
        utterance_constraints = self._extract_utterance_constraints(sections.get("constraints", ""))
        flow_nodes = self._parse_flow_from_text(sections.get("call_flow", ""))
        knowledge_base = self._parse_knowledge_from_text(sections.get("knowledge", ""))
        variables = self._extract_variables(markdown_text)
        scenario_triggers = self._extract_scenario_triggers(sections.get("constraints", ""))
        exception_rules = self._extract_exception_rules(sections)
        # 规则模式下, 检查点从流程派生 (在 compiler 中会标记 derived_from_flow=True)
        checkpoints = self._derive_checkpoints_from_flow(flow_nodes)

        system_prompt = self._build_system_prompt(role_desc, task_desc, sections)
        task_name = self._infer_task_name(role_desc, task_desc)

        return TaskInstruction(
            task_name=task_name,
            task_description=task_desc or role_desc,
            system_prompt=system_prompt,
            role_description=role_desc,
            opening_line=opening_line,
            checkpoints=checkpoints,
            constraints=constraints,
            flow_nodes=flow_nodes,
            exception_rules=exception_rules,
            knowledge_base=knowledge_base,
            variables=variables,
            utterance_constraints=utterance_constraints,
            scenario_triggers=scenario_triggers,
            instruction_format="markdown",
            raw_instruction=markdown_text,
        )

    # ---- Markdown 切段 ----

    def _extract_markdown_sections(self, text: str) -> dict[str, str]:
        sections: dict[str, str] = {}
        # 关键: 切段时只切 # 一级标题, 二级标题(## Step, ## Task, ## Constraints) 内容应保留在所属 section 中
        # 但有些样本(xlsx task2)只用 # Role 一级 + ## Task 二级, 此时需要二级也作为 section 入口
        # 解决: 第一遍按 # 切, 然后把每个 section 内的 ## H2 提升到顶层(只有特定关键字)
        parts = re.split(r"\n(?=#\s)", text)
        if not text.lstrip().startswith("#"):
            sections["preamble"] = parts[0].strip()
            parts = parts[1:]

        # 关键 H2 引子: 这些二级标题如果出现, 也作为 section 起点
        h2_promotable = ["task", "constraints", "constraint", "conversation flow", "call flow",
                         "knowledge", "faq", "opening line", "role"]

        for part in parts:
            # 每个 part 里, 检查是否包含 promotable 的 ## 标题
            # 用 ## 进一步切
            sub_parts = re.split(r"\n(?=##\s)", part)
            for sp in sub_parts:
                self._classify_section(sp.strip(), sections, h2_promotable)

        return sections

    @staticmethod
    def _classify_section(part: str, sections: dict[str, str], h2_promotable: list[str]):
        """分类一个段落到 sections 字典"""
        if not part:
            return
        lines = part.split("\n", 1)
        if not lines:
            return
        header_raw = lines[0].strip()
        # 去掉所有前导 #
        header = header_raw.lstrip("#").strip()
        # 去掉冒号
        if ":" in header:
            head_part, _, _ = header.partition(":")
        elif "：" in header:
            head_part, _, _ = header.partition("：")
        else:
            head_part = header
        head_part = head_part.strip()
        head_lower = head_part.lower()

        content = lines[1].strip() if len(lines) > 1 else ""

        # 如果标题里含冒号, 冒号后的内容也算 content
        if ":" in header or "：" in header:
            sep = ":" if ":" in header else "："
            after_colon = header.split(sep, 1)[1].strip()
            if after_colon and not content:
                content = after_colon
            elif after_colon and content:
                content = after_colon + "\n" + content

        # 分类
        if "role" in head_lower and "task" not in head_lower:
            sections["role"] = content or head_part
        elif "task" in head_lower:
            sections["task"] = content
        elif "opening" in head_lower:
            sections["opening_line"] = content
        elif "flow" in head_lower or "conversation" in head_lower:
            # 注意: ## Step N 也会被这个匹配吃, 但 step 不应作为 section 入口
            if not re.match(r"^step\s*\d", head_lower):
                sections["call_flow"] = sections.get("call_flow", "") + ("\n" if sections.get("call_flow") else "") + part.strip()
        elif "knowledge" in head_lower or "faq" in head_lower:
            sections["knowledge"] = content
        elif "constraint" in head_lower:
            sections["constraints"] = content
        elif re.match(r"^step\s*\d", head_lower):
            # ## Step N: ... 应当合并到 call_flow
            sections["call_flow"] = sections.get("call_flow", "") + ("\n\n" if sections.get("call_flow") else "") + part.strip()
        else:
            sections[head_lower] = content

    # ---- 规则抽取 (与 v2 类似但增强 normalisation) ----

    def _parse_constraints_from_text(self, text: str) -> list[Constraint]:
        constraints: list[Constraint] = []
        if not text:
            return constraints
        items = re.findall(r"[-*]\s*(.+?)(?=\n[-*]|\Z)", text, re.DOTALL)
        if not items:
            items = [line.strip() for line in text.split("\n") if line.strip() and not line.startswith("#")]

        for item in items:
            item = item.strip()
            if not item:
                continue
            severity = ConstraintSeverity.MAJOR
            if any(kw in item for kw in ["不能", "禁止", "不可", "严禁", "不允许"]):
                severity = ConstraintSeverity.CRITICAL
            elif any(kw in item for kw in ["避免", "尽量不", "减少"]):
                severity = ConstraintSeverity.MINOR

            check_method = "llm_judge"
            if re.search(r"\d+\s*[个字]", item):
                check_method = "char_count"

            constraints.append(Constraint(
                name=(item[:20] + ("..." if len(item) > 20 else "")),
                description=item,
                severity=severity,
                check_method=check_method,
            ))
        return constraints

    def _extract_utterance_constraints(self, text: str) -> list[UtteranceConstraint]:
        ucs: list[UtteranceConstraint] = []
        if not text:
            return ucs

        # 字数 (兼容: "最多15-20个字" / "30个字以内" / "约30个字")
        char_re = re.compile(
            r"(?:每次回复|单句|每[句条])[^\n]{0,15}?(?:最多|不超过|控制在|≤|约).{0,5}?"
            r"(\d+)(?:\s*[-~到至]\s*(\d+))?\s*[个]?[字字符]"
        )
        m = char_re.search(text)
        if not m:
            m = re.search(r"(\d+)\s*[-~到至]?\s*(\d*)\s*[个]?[字字符](?:以内|左右|之内)", text)
        if m:
            high = int(m.group(2)) if m.group(2) else int(m.group(1))
            ucs.append(UtteranceConstraint(
                name=f"单句字数限制({high}字)",
                max_chars=high,
                description=m.group(0),
            ))

        # 禁词 (双引号 / 中文引号 / 普通引号 / "等语气词")
        forbidden_set: set[str] = set()
        for pat in [
            r'不[说用]"([^"]{1,15})"',
            r'不[说用]"([^"]{1,15})"',
            r"不[说用]'([^']{1,15})'",
            r"不[说用]([^\n，。、,；]{0,40}?)等[语气词词汇表达]",
            r"避免[说用]([^\n，。]{0,40}?)等",
        ]:
            for raw in re.findall(pat, text):
                # 多词分割
                for w in re.split(r"[、，,/\s]+", raw):
                    w = w.strip().strip('"').strip('"').strip("'")
                    if w and len(w) <= 8:
                        forbidden_set.add(w)
        if forbidden_set:
            ucs.append(UtteranceConstraint(
                name="禁止语气词/词汇",
                forbidden_words=sorted(forbidden_set),
                description="禁止使用: " + ", ".join(sorted(forbidden_set)),
            ))

        # 语气
        m = re.search(r"(?:保持|语气).{0,8}?(随意|自然|简洁|正式|专业|友好|口语化)", text)
        if m:
            ucs.append(UtteranceConstraint(
                name="语气风格要求",
                required_tone=m.group(1),
                description=m.group(0),
            ))

        return ucs

    def _parse_flow_from_text(self, text: str) -> list[FlowNode]:
        nodes: list[FlowNode] = []
        if not text:
            return nodes

        # 多种步骤格式
        # ## Step N: title  /  N. title  /  ### N.M
        block_pattern = re.compile(
            r"(?:##\s*Step\s*(\d+(?:\.\d+)?)[:：\s]+(.+?)(?=\n##|\Z))|"
            r"(?:^(\d+)[.、)]\s+(.+?)(?=\n\d+[.、)]|\Z))",
            re.MULTILINE | re.DOTALL,
        )
        steps: list[tuple[str, str]] = []
        for m in block_pattern.finditer(text):
            if m.group(1):
                steps.append((m.group(1), m.group(2).strip()))
            elif m.group(3):
                steps.append((m.group(3), m.group(4).strip()))

        for i, (step_num, content) in enumerate(steps):
            branches = self._extract_branches_from_text(content)
            next_nodes = []
            if i < len(steps) - 1:
                next_nodes.append(f"步骤{steps[i + 1][0]}")
            script_match = re.search(r"(?:参考话术|参考)[^\n]*?[：:]\s*(.+?)(?=\n[-*#]|\Z)", content)
            ref_script = script_match.group(1).strip() if script_match else ""
            sub_steps = re.findall(r"(?:^\s*\d+\.|\s*[-*])\s+(.+)", content)

            try:
                step_no_int = int(float(step_num))
            except ValueError:
                step_no_int = i + 1

            nodes.append(FlowNode(
                name=f"步骤{step_num}",
                description=content.split("\n")[0].strip()[:200],
                step_number=step_no_int,
                next_nodes=next_nodes,
                is_start=(i == 0),
                is_end=(i == len(steps) - 1),
                branches=branches,
                reference_script=ref_script,
                sub_steps=sub_steps[:10],
            ))
        return nodes

    def _extract_branches_from_text(self, text: str) -> list[ConditionalBranch]:
        branches: list[ConditionalBranch] = []
        # 中英文兼容的条件分支匹配 - 顺序敏感, 优先级高的先匹配
        patterns = [
            r"[-*]\s*若(.+?)[→:：]\s*(.+?)(?=\n[-*]|\Z)",
            r"[-*]\s*如果(.+?)[→:：]\s*(.+?)(?=\n[-*]|\Z)",
            r"[-*]\s*(?:if|If|IF)\s+(.+?)\s*[:→]\s*(.+?)(?=\n[-*]|\Z)",
        ]
        seen_norm: set[str] = set()
        for pat in patterns:
            for cond, action in re.findall(pat, text, re.DOTALL):
                cond = cond.strip().rstrip("，,。.")
                action = action.strip()
                if not cond or not action:
                    continue
                # 归一化: 去掉 "若是"/"是"/"若" 前缀差异
                norm = re.sub(r"^(若是|若|如果是|如果)", "", cond).strip()
                if norm in seen_norm:
                    continue
                seen_norm.add(norm)
                branches.append(ConditionalBranch(
                    condition=cond,
                    action=action,
                    is_default=any(kw in cond for kw in ["否则", "其他", "默认", "否则的话"]),
                ))
        return branches

    def _parse_knowledge_from_text(self, text: str) -> list[KnowledgeEntry]:
        entries: list[KnowledgeEntry] = []
        if not text:
            return entries
        items = re.findall(r"[-*]\s*(.+?)(?=\n[-*]|\Z)", text, re.DOTALL)
        if not items:
            items = [line.strip() for line in text.split("\n") if line.strip()]
        for item in items:
            item = item.strip()
            if not item or item.startswith("#"):
                continue
            qa = re.match(r"(.+?)[：:]\s*(.+)", item, re.DOTALL)
            if qa:
                question = qa.group(1).strip()[:50]
                answer = qa.group(2).strip()
            else:
                question = item[:30]
                answer = item
            keywords = re.findall(r"\*\*(.+?)\*\*", item)
            entries.append(KnowledgeEntry(
                question=question,
                answer=answer,
                keywords=keywords,
            ))
        return entries

    def _extract_variables(self, text: str) -> list[VariableDefinition]:
        vars_: list[VariableDefinition] = []
        seen: set[str] = set()
        # ${name}
        for vname in re.findall(r"\$\{([^}]+)\}", text):
            if vname in seen:
                continue
            seen.add(vname)
            vars_.append(VariableDefinition(
                name=vname,
                placeholder=f"${{{vname}}}",
                description=f"动态变量: {vname}",
                default_value=self._guess_default(vname),
            ))
        # **X 单**
        for m in re.finditer(r"\*\*([A-Z])\s*([单天元个次分时])\*\*", text):
            vname = f"{m.group(1)}_{m.group(2)}"
            if vname in seen:
                continue
            seen.add(vname)
            vars_.append(VariableDefinition(
                name=vname,
                placeholder=m.group(0),
                description=f"数值变量: {m.group(1)} ({m.group(2)})",
                sample_values=["3", "5", "10"],
                default_value="5",
            ))
        return vars_

    @staticmethod
    def _guess_default(var_name: str) -> str:
        name = var_name.lower()
        if "name" in name:
            return "张师傅"
        if "phone" in name or "tel" in name:
            return "138****8888"
        if "addr" in name or "address" in name:
            return "北京市朝阳区建国路88号"
        if "order" in name:
            return "8856"
        if "time" in name or "date" in name:
            return "今天14:00-16:00"
        return ""

    def _extract_scenario_triggers(self, text: str) -> list[ScenarioTrigger]:
        triggers: list[ScenarioTrigger] = []
        if not text:
            return triggers
        # 通用模式
        for cond, quote, action in re.findall(r"[-*]\s*若(.+?)(?:说|表示)[\"\"]([^\"\"]+)[\"\"](?:后)?(.+?)(?=\n[-*]|\Z)", text):
            should_end = any(kw in action for kw in ["挂断", "结束", "再打", "再联系"])
            triggers.append(ScenarioTrigger(
                trigger_condition=cond.strip(),
                trigger_keywords=[quote.strip()],
                expected_system_action=f"{quote.strip()} {action.strip()}",
                should_end_conversation=should_end,
            ))
        # 特定场景
        for pat, kws, action, end in [
            (r"(?:在开车|开车中)", ["在开车", "开车"], "礼貌告知稍后再联系并结束通话", True),
            (r"(?:很忙|忙着|没时间|在忙)", ["很忙", "忙", "没时间", "在忙"], "简短说明要点或约定稍后再联系", False),
            (r"(?:不需要|不感兴趣|不用了)", ["不需要", "不感兴趣", "不用"], "尊重用户意愿，礼貌结束", True),
        ]:
            if re.search(pat, text):
                triggers.append(ScenarioTrigger(
                    trigger_condition=f"用户表示{kws[0]}",
                    trigger_keywords=kws,
                    expected_system_action=action,
                    should_end_conversation=end,
                ))
        return triggers

    def _derive_checkpoints_from_flow(self, flow_nodes: list[FlowNode]) -> list[Checkpoint]:
        cps: list[Checkpoint] = []
        for node in flow_nodes:
            cp_type = CheckpointType.MUST_DO
            if node.is_end:
                cp_type = CheckpointType.SHOULD_DO
            cps.append(Checkpoint(
                name=node.name,
                description=node.description,
                type=cp_type,
                order=node.step_number,
                at_flow_step=node.name,
                evaluation_criteria=f"完成{node.name}所要求的对话动作",
                derived_from_flow=True,  # 关键: 标记派生
            ))
        return cps

    def _extract_exception_rules(self, sections: dict[str, str]) -> list[ExceptionRule]:
        rules: list[ExceptionRule] = []
        text = sections.get("constraints", "")
        for pat, trigger in [
            (r'如(?:被问|遇到).*?超出.*?[：:]?\s*[\""](.+?)[\""]', "被问及超出职责范围的问题"),
            (r"如果.*?坚持.*?(?:无法|不能).*?[，,]\s*(.+?)(?=\n|$)", "用户坚持无法配合"),
        ]:
            m = re.search(pat, text)
            if m:
                rules.append(ExceptionRule(trigger=trigger, expected_behavior=m.group(1).strip()))
        return rules

    def _build_system_prompt(self, role_desc: str, task_desc: str, sections: dict[str, str]) -> str:
        parts = []
        if role_desc:
            parts.append(f"## 角色\n{role_desc}")
        if task_desc:
            parts.append(f"## 任务\n{task_desc}")
        if sections.get("knowledge"):
            parts.append(f"## 知识库\n{sections['knowledge']}")
        if sections.get("constraints"):
            parts.append(f"## 约束规则\n{sections['constraints']}")
        if sections.get("call_flow"):
            parts.append(f"## 对话流程\n{sections['call_flow']}")
        return "\n\n".join(parts)

    @staticmethod
    def _infer_task_name(role_desc: str, task_desc: str) -> str:
        if task_desc:
            # 清理 markdown 标题前缀
            t = task_desc.strip()
            t = re.sub(r"^#+\s*", "", t)
            t = re.sub(r"^Task[:：]\s*", "", t, flags=re.IGNORECASE)
            t = re.sub(r"^任务[:：]\s*", "", t)
            first = t.split("\n")[0].split("。")[0].strip()
            return (first[:30] + "...") if len(first) > 30 else first
        if role_desc:
            r = re.sub(r"^Role[:：]\s*", "", role_desc, flags=re.IGNORECASE).strip()
            return r.split("\n")[0][:30]
        return "未命名任务"

    # ---- 结构化数据装配 ----

    def _parse_structured_data(self, data: dict[str, Any]) -> TaskInstruction:
        # 检查点
        checkpoints = []
        for cp in data.get("checkpoints", []):
            try:
                cp_type = CheckpointType(cp.get("type", "must_do"))
            except ValueError:
                cp_type = CheckpointType.MUST_DO
            checkpoints.append(Checkpoint(
                name=cp.get("name", ""),
                description=cp.get("description", ""),
                type=cp_type,
                keywords=cp.get("keywords", []),
                evaluation_criteria=cp.get("evaluation_criteria", ""),
                order=int(cp.get("order", 0) or 0),
                condition=cp.get("condition", ""),
                at_flow_step=cp.get("at_flow_step", ""),
                derived_from_flow=cp.get("derived_from_flow", False),
            ))

        # 约束
        constraints = []
        for c in data.get("constraints", []):
            try:
                sev = ConstraintSeverity(c.get("severity", "major"))
            except ValueError:
                sev = ConstraintSeverity.MAJOR
            constraints.append(Constraint(
                name=c.get("name", ""),
                description=c.get("description", ""),
                severity=sev,
                violation_examples=c.get("violation_examples", []),
                check_method=c.get("check_method", "llm_judge"),
            ))

        # 流程节点
        flow_nodes = []
        for fn in data.get("flow_nodes", []):
            branches = []
            for br in fn.get("branches", []):
                branches.append(ConditionalBranch(
                    condition=br.get("condition", ""),
                    condition_keywords=br.get("condition_keywords", []),
                    action=br.get("action", ""),
                    next_node=br.get("next_node", ""),
                    is_default=br.get("is_default", False),
                ))
            try:
                step_no = int(float(fn.get("step_number", 0) or 0))
            except (ValueError, TypeError):
                step_no = 0
            flow_nodes.append(FlowNode(
                name=fn.get("name", ""),
                description=fn.get("description", ""),
                next_nodes=fn.get("next_nodes", []),
                is_start=fn.get("is_start", False),
                is_end=fn.get("is_end", False),
                conditions=fn.get("conditions", {}),
                branches=branches,
                step_number=step_no,
                reference_script=fn.get("reference_script", ""),
                sub_steps=fn.get("sub_steps", []),
            ))

        exception_rules = [
            ExceptionRule(
                trigger=er.get("trigger", ""),
                expected_behavior=er.get("expected_behavior", ""),
                forbidden_behavior=er.get("forbidden_behavior", ""),
            )
            for er in data.get("exception_rules", [])
        ]

        knowledge_base = [
            KnowledgeEntry(
                question=kb.get("question", ""),
                answer=kb.get("answer", ""),
                keywords=kb.get("keywords", []),
                category=kb.get("category", "general"),
                priority=int(kb.get("priority", 0) or 0),
            )
            for kb in data.get("knowledge_base", [])
        ]

        variables = [
            VariableDefinition(
                name=v.get("name", ""),
                description=v.get("description", ""),
                placeholder=v.get("placeholder", ""),
                sample_values=v.get("sample_values", []),
                default_value=v.get("default_value", ""),
            )
            for v in data.get("variables", [])
        ]

        utterance_constraints = [
            UtteranceConstraint(
                name=uc.get("name", ""),
                max_chars=uc.get("max_chars"),
                min_chars=uc.get("min_chars"),
                forbidden_words=uc.get("forbidden_words", []),
                required_tone=uc.get("required_tone", ""),
                applies_to=uc.get("applies_to", "all"),
                description=uc.get("description", ""),
            )
            for uc in data.get("utterance_constraints", [])
        ]

        scenario_triggers = [
            ScenarioTrigger(
                trigger_condition=st.get("trigger_condition", ""),
                trigger_keywords=st.get("trigger_keywords", []),
                expected_system_action=st.get("expected_system_action", ""),
                forbidden_system_action=st.get("forbidden_system_action", ""),
                should_end_conversation=st.get("should_end_conversation", False),
                at_flow_step=st.get("at_flow_step", ""),
            )
            for st in data.get("scenario_triggers", [])
        ]

        instr = TaskInstruction(
            task_name=data.get("task_name", "未命名任务"),
            task_description=data.get("task_description", ""),
            system_prompt=data.get("system_prompt", ""),
            context_info=data.get("context_info", {}),
            checkpoints=checkpoints,
            constraints=constraints,
            flow_nodes=flow_nodes,
            exception_rules=exception_rules,
            opening_line=data.get("opening_line", ""),
            knowledge_base=knowledge_base,
            variables=variables,
            utterance_constraints=utterance_constraints,
            scenario_triggers=scenario_triggers,
            role_description=data.get("role_description", ""),
            instruction_format=data.get("instruction_format", "json"),
        )

        logger.info(
            f"指令解析完成: {instr.task_name} | CP={len(checkpoints)} CON={len(constraints)} "
            f"FN={len(flow_nodes)} KB={len(knowledge_base)} UC={len(utterance_constraints)} "
            f"ST={len(scenario_triggers)}"
        )
        return instr

    @staticmethod
    def _parse_json(text: str) -> dict[str, Any]:
        if "```json" in text:
            start = text.find("```json") + 7
            end = text.find("```", start)
            if end != -1:
                text = text[start:end].strip()
        elif "```" in text:
            start = text.find("```") + 3
            end = text.find("```", start)
            if end != -1:
                text = text[start:end].strip()
        if not text.lstrip().startswith("{"):
            s = text.find("{")
            e = text.rfind("}")
            if s != -1 and e != -1:
                text = text[s:e + 1]
        try:
            return json.loads(text)
        except json.JSONDecodeError as e:
            logger.warning(f"JSON parse failed: {e}; head={text[:120]!r}")
            return {}

    # ---- 验证 ----

    def validate_instruction(self, instruction: TaskInstruction) -> list[str]:
        warnings: list[str] = []
        if not instruction.checkpoints:
            warnings.append("警告: 任务指令中没有定义任何检查点, 将无法评估任务完成度")
        if not instruction.system_prompt:
            warnings.append("警告: 未设置系统提示词")
        if not instruction.opening_line:
            warnings.append("提示: 未设置开场白, 将由模型自动生成")

        must_do = sum(1 for cp in instruction.checkpoints if cp.type == CheckpointType.MUST_DO)
        if instruction.checkpoints and must_do == 0:
            warnings.append("提示: 没有定义必达检查点 (must_do)")

        if instruction.flow_nodes:
            if not any(n.is_start for n in instruction.flow_nodes):
                warnings.append("警告: 流程图中没有定义起始节点")
            if not any(n.is_end for n in instruction.flow_nodes):
                warnings.append("警告: 流程图中没有定义结束节点")

        for uc in instruction.utterance_constraints:
            if uc.max_chars is not None and uc.max_chars < 5:
                warnings.append(f"警告: 话术约束 '{uc.name}' 字数限制过小 ({uc.max_chars})")

        for kb in instruction.knowledge_base:
            if not kb.answer:
                warnings.append(f"警告: 知识库条目 '{kb.question}' 缺少答案")

        # 变量检查
        if instruction.variables and instruction.opening_line:
            for v in instruction.variables:
                if v.placeholder and v.placeholder in instruction.opening_line:
                    if not v.default_value and not v.sample_values:
                        warnings.append(f"提示: 开场白包含变量 '{v.placeholder}' 但未设默认值")

        return warnings
